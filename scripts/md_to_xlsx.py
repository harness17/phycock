"""
md_to_xlsx.py - Markdown → Excel 汎用変換スクリプト

使い方:
  python scripts/md_to_xlsx.py <入力.md> <出力.xlsx>

動作:
  - H2 セクション（## ）ごとに1シートを作成する
  - H3 見出し（### ）はシート内のセクションタイトル行として出力する
  - Markdown テーブル（| col | col |）は自動的にスタイル付きテーブルに変換する
  - テーブル以外のテキスト行は薄いグレーで出力する（コンテキスト保持）
  - H1（# ）はシート名には使わず、最初のシートのタイトル行として出力する
"""

import re
import sys
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─────────────────────────────────────────────────────────────
# カラーパレット
# ─────────────────────────────────────────────────────────────
COLOR = {
    'header_bg':  '2E75B6',   # テーブルヘッダー（青）
    'header_fg':  'FFFFFF',   # テーブルヘッダー文字（白）
    'alt_bg':     'DEEAF1',   # 偶数行（薄青）
    'h1_bg':      '1F4E79',   # H1タイトル（濃紺）
    'h1_fg':      'FFFFFF',
    'h2_bg':      '2E75B6',   # H2セクション（青）
    'h2_fg':      'FFFFFF',
    'h3_bg':      'BDD7EE',   # H3サブセクション（薄青）
    'h3_fg':      '1F4E79',
    'text_bg':    'F5F5F5',   # 通常テキスト行（薄グレー）
    'border':     'AAAAAA',
}

def fill(hex_color):
    return PatternFill(fill_type='solid', fgColor=hex_color)

def font(bold=False, color='000000', size=10):
    return Font(bold=bold, color=color, size=size, name='Meiryo UI')

def border():
    s = Side(style='thin', color=COLOR['border'])
    return Border(left=s, right=s, top=s, bottom=s)

def align(wrap=True):
    return Alignment(wrap_text=wrap, vertical='center')

def set_cell(ws, row, col, value, bg=None, fg='000000', bold=False, size=10, merge_to=None):
    """セルに値・スタイルを設定する"""
    cell = ws.cell(row=row, column=col, value=str(value) if value is not None else '')
    cell.font = font(bold=bold, color=fg, size=size)
    cell.alignment = align()
    cell.border = border()
    if bg:
        cell.fill = fill(bg)
    return cell

# ─────────────────────────────────────────────────────────────
# Markdown パーサー
# ─────────────────────────────────────────────────────────────

def is_table_row(line):
    """テーブル行（| で始まる）かどうか判定"""
    return line.strip().startswith('|') and line.strip().endswith('|')

def is_separator_row(line):
    """テーブルセパレーター行（|---|---| 形式）かどうか判定"""
    return is_table_row(line) and re.match(r'^\|[\s\-:|]+\|', line.strip())

def parse_table_row(line):
    """テーブル行のセル値をリストで返す"""
    cells = line.strip().strip('|').split('|')
    return [c.strip() for c in cells]

def clean_markdown(text):
    """Markdown の修飾記号を除去してプレーンテキストにする"""
    text = re.sub(r'\*\*(.*?)\*\*', r'\1', text)   # **bold**
    text = re.sub(r'\*(.*?)\*', r'\1', text)         # *italic*
    text = re.sub(r'`(.*?)`', r'\1', text)           # `code`
    text = re.sub(r'\[([^\]]+)\]\([^\)]+\)', r'\1', text)  # [text](url)
    return text.strip()

def parse_sections(lines):
    """
    Markdown を H2 セクション単位でパースする。
    各セクションは { 'title': str, 'blocks': list } の形式。
    blocks の各要素は:
      { 'type': 'h1'|'h2'|'h3'|'text'|'table', ... }
    """
    sections = []
    current_section = {'title': '概要', 'blocks': []}
    i = 0
    while i < len(lines):
        line = lines[i]
        stripped = line.strip()

        # H1 → 最初のセクションのタイトルブロック
        if re.match(r'^# ', stripped):
            current_section['blocks'].append({
                'type': 'h1',
                'text': clean_markdown(stripped[2:])
            })

        # H2 → 新しいシートを開始
        elif re.match(r'^## ', stripped):
            if current_section['blocks']:
                sections.append(current_section)
            current_section = {
                'title': clean_markdown(stripped[3:]),
                'blocks': []
            }

        # H3 → セクション内の見出し
        elif re.match(r'^### ', stripped):
            current_section['blocks'].append({
                'type': 'h3',
                'text': clean_markdown(stripped[4:])
            })

        # テーブル → ヘッダー行・セパレーター・データ行をまとめて収集
        elif is_table_row(stripped):
            rows = []
            while i < len(lines) and is_table_row(lines[i].strip()):
                if not is_separator_row(lines[i]):
                    rows.append(parse_table_row(lines[i]))
                i += 1
            if rows:
                current_section['blocks'].append({
                    'type': 'table',
                    'headers': rows[0],
                    'rows': rows[1:],
                })
            continue  # i は既に進んでいるのでインクリメントしない

        # 水平線・空行・目次リンクはスキップ
        elif stripped in ('---', '') or re.match(r'^\d+\. \[', stripped):
            pass

        # コードブロックはスキップ（開始 ``` が来たら対応する ``` まで読み飛ばす）
        elif stripped.startswith('```'):
            i += 1
            while i < len(lines) and not lines[i].strip().startswith('```'):
                i += 1

        # 通常テキスト
        elif stripped:
            current_section['blocks'].append({
                'type': 'text',
                'text': clean_markdown(stripped)
            })

        i += 1

    if current_section['blocks']:
        sections.append(current_section)

    return sections

# ─────────────────────────────────────────────────────────────
# Excel 書き出し
# ─────────────────────────────────────────────────────────────

def write_section_to_sheet(wb, section, max_col_width=60):
    """1セクション = 1シートとして書き出す"""
    # Excel のシート名は最大 31 文字、使えない文字あり
    raw_title = section['title']
    sheet_name = re.sub(r'[\\/*?\[\]:]', '', raw_title)[:31]
    ws = wb.create_sheet(title=sheet_name)

    row = 1
    col_widths = {}  # 列ごとの最大文字幅を追跡

    for block in section['blocks']:

        # H1 タイトル行
        if block['type'] == 'h1':
            set_cell(ws, row, 1, block['text'], bg=COLOR['h1_bg'], fg=COLOR['h1_fg'], bold=True, size=14)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.row_dimensions[row].height = 24
            row += 1

        # H3 サブセクション行
        elif block['type'] == 'h3':
            row += 1  # 前のブロックとの余白
            set_cell(ws, row, 1, block['text'], bg=COLOR['h3_bg'], fg=COLOR['h3_fg'], bold=True, size=11)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.row_dimensions[row].height = 18
            row += 1

        # テキスト行
        elif block['type'] == 'text':
            set_cell(ws, row, 1, block['text'], bg=COLOR['text_bg'])
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.row_dimensions[row].height = 15
            row += 1

        # テーブル
        elif block['type'] == 'table':
            headers = block['headers']
            n_cols = len(headers)

            # ヘッダー行
            for c, h in enumerate(headers, 1):
                set_cell(ws, row, c, h, bg=COLOR['header_bg'], fg=COLOR['header_fg'], bold=True, size=10)
                col_widths[c] = max(col_widths.get(c, 0), min(len(h) * 2.2, max_col_width))
            ws.row_dimensions[row].height = 16
            row += 1

            # データ行
            for di, data_row in enumerate(block['rows']):
                bg = COLOR['alt_bg'] if di % 2 == 1 else None
                for c, val in enumerate(data_row, 1):
                    set_cell(ws, row, c, val, bg=bg)
                    col_widths[c] = max(col_widths.get(c, 0), min(len(val) * 1.8, max_col_width))
                # 列数が足りない場合は空白セルで埋める
                for c in range(len(data_row) + 1, n_cols + 1):
                    set_cell(ws, row, c, '', bg=bg)
                ws.row_dimensions[row].height = 15
                row += 1

            row += 1  # テーブル後の余白

    # 列幅を設定
    for col_idx, width in col_widths.items():
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(width, 8)

    # ウィンドウ枠の固定（テーブルが多い場合に先頭行を固定）
    ws.freeze_panes = 'A2'

    return ws

# ─────────────────────────────────────────────────────────────
# メイン処理
# ─────────────────────────────────────────────────────────────

def convert(md_path, xlsx_path):
    print(f'変換開始: {md_path} → {xlsx_path}')

    with open(md_path, encoding='utf-8') as f:
        lines = f.readlines()

    sections = parse_sections(lines)
    print(f'  セクション数: {len(sections)}')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # デフォルトシートを削除

    for section in sections:
        ws = write_section_to_sheet(wb, section)
        n_rows = ws.max_row
        print(f'  シート作成: [{ws.title}] ({n_rows} 行)')

    wb.save(xlsx_path)
    size = os.path.getsize(xlsx_path)
    print(f'完了: {xlsx_path} ({size:,} bytes)')

if __name__ == '__main__':
    if len(sys.argv) != 3:
        print('使い方: python scripts/md_to_xlsx.py <入力.md> <出力.xlsx>')
        sys.exit(1)

    md_path  = sys.argv[1]
    xlsx_path = sys.argv[2]

    if not os.path.exists(md_path):
        print(f'エラー: ファイルが見つかりません: {md_path}')
        sys.exit(1)

    convert(md_path, xlsx_path)
